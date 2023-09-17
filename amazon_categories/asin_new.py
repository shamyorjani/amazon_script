import openpyxl
import os
import pyperclip
import requests
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Define the path to the CRX file
extension_path = './amazoncrxextension.crx'

# Set up Chrome WebDriver with the extension
chrome_options = webdriver.ChromeOptions()
chrome_options.add_extension(extension_path)

# Create the WebDriver with ChromeOptions
driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()

# Load URLs from "amazon_categories_all.xlsx" file
excel_file_path = './amazon_categories_all.xlsx'
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# Extract URLs from column B (URLs)
amazon_urls = [cell.value for cell in sheet['B'] if cell.value and cell.row > 1]

# Set up WebDriverWait
wait = WebDriverWait(driver, 10)

# Create a new workbook or load an existing one
excel_file_path = './ASINs.xlsx'

if os.path.isfile(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
    worksheet = workbook.active
else:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(1, 1).value = 'ASINs'
    worksheet.cell(1, 2).value = 'Category'  # Add a new column for category names
    worksheet.cell(1, 1).font = openpyxl.styles.Font(bold=True)
    worksheet.cell(1, 2).font = openpyxl.styles.Font(bold=True)

# Get the last row in the Excel sheet
last_row = worksheet.max_row + 1

# Loop through each URL
for url in amazon_urls:
    driver.get(url)
    time.sleep(3)

    try:
        # Check for the presence of the ASIN Extractor dropdown
        asin_extractor = wait.until(EC.element_to_be_clickable((By.ID, 'azASINExtractorDropDown')))
        asin_extractor.click()

        # Check for the presence of the "All ASINs (100)" option in the dropdown
        all_asins_option = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'header') and contains(text(), 'All ASINs')]")))
        all_asins_option.click()

        time.sleep(3)
        # Paste the captured ASINs text from the clipboard
        captured_asins = pyperclip.paste()

        # Split the pasted text into ASINs and append them to the Excel file
        asins_list = captured_asins.split("\n")

        # Extract the category name from the specified h1 element
        category_name = driver.find_element(By.XPATH, "//h1[contains(@class, 'a-size-large')]").text
        # Remove "Best Sellers in" from the category name
        category_name = category_name.replace("Best Sellers in ", "")

        for value in asins_list:
            cell_asin = worksheet.cell(row=last_row, column=1)
            cell_asin.value = value
            
            # Store the modified category name for all ASINs
            cell_category = worksheet.cell(row=last_row, column=2)  # New column for category names
            cell_category.value = category_name
            
            last_row += 1

    except Exception as e:
        pass

# Remove duplicate ASIN rows
unique_asins = set()
rows_to_delete = []
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=1):
    for cell in row:
        if cell.value in unique_asins:
            rows_to_delete.append(cell.row)
        else:
            unique_asins.add(cell.value)

for row_number in reversed(rows_to_delete):
    worksheet.delete_rows(row_number)

# Save the Excel workbook
workbook.save(excel_file_path)

# Close the WebDriver
driver.quit()

print(f"ASINs and category name have been captured, modified, and duplicate ASIN rows removed. Data saved to {excel_file_path}")
