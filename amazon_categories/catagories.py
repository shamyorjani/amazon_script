from time import sleep
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl

# Define the path to the CRX file
extension_path = './amazoncrxextension.crx'

# Set up Chrome WebDriver with the extension
chrome_options = webdriver.ChromeOptions()
chrome_options.add_extension(extension_path)

# Create the WebDriver with ChromeOptions
driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()

# Switch to the new tab
driver.switch_to.window(driver.window_handles[0])

# Define the base URL of Amazon.de
base_url = 'https://www.amazon.de/'
url = []
# Create a single Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# get main categories list
amazon_url = "https://www.amazon.de/-/en/gp/bestsellers/?ref_=nav_cs_bestsellers"
sleep(15)
response = driver.get(amazon_url)

wait = WebDriverWait(driver, 5)
try:
    # Check for the presence of the "sp-cc-rejectall-link" element
    reject_button = wait.until(
        EC.element_to_be_clickable((By.ID, 'sp-cc-rejectall-link')))

    # Click the "Continue without accepting" link
    reject_button.click()
except Exception as e:
    pass
# Parse the HTML content of the page
sleep(10)
soup = BeautifulSoup(driver.page_source, "lxml")

# Find all div elements with role="treeitem" and class="_p13n-zg-nav-tree-all_style_zg-browse-item__1rdKf _p13n-zg-nav-tree-all_style_zg-browse-height-small__nleKL"
category_divs = soup.find_all("div", {
                              "role": "treeitem", "class": "_p13n-zg-nav-tree-all_style_zg-browse-item__1rdKf _p13n-zg-nav-tree-all_style_zg-browse-height-small__nleKL"})
# Loop through the category divs
# Loop through the category divs
for category_div in category_divs:
    # Try to find the <a> element within the category div to get the URL
    category_link = category_div.find('a')

    if category_link:
        # Extract the href attribute (URL)
        category_url = category_link.get("href")

        # Check if the URL starts with "/"
        if category_url.startswith("/"):
            # Complete the URL by adding the Amazon domain
            category_url = "https://www.amazon.de" + category_url

            # Extract the category name
            category_name = category_link.text.strip()

            # Append the category URL and category name to the list
            url.append({"name": category_name, "url": category_url})

# Print the list of URLs and category names
# for item in url:
#     print(f"Name: {item['name']}")
#     print(f"URL: {item['url']}")

max_retries = 3
# Define an array of URLs and their names
url_data = []
sub_url_data = []

def main_scrape_data_from_url(urls):
    row = 1
    for url in urls:
        for _ in range(max_retries):
            try:
                # Navigate to the URL using Selenium
                driver.get(url['url'])

                # Wait for the page to load and find the relevant elements
                WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                    (By.CLASS_NAME, '_p13n-zg-nav-tree-all_style_zg-browse-group__88fbz')))

                # Parse the HTML content of the page using BeautifulSoup
                soup = BeautifulSoup(driver.page_source, 'html.parser')

                # Find all div elements with role="group" and class="_p13n-zg-nav-tree-all_style_zg-browse-group__88fbz"
                group_divs = soup.find_all('div', {
                    'role': 'group', 'class': '_p13n-zg-nav-tree-all_style_zg-browse-group__88fbz'})

                # Loop through each group div and extract names and URLs
                for group_div in group_divs:
                    # Find all div elements with role="treeitem" and class="_p13n-zg-nav-tree-all_style_zg-browse-item__1rdKf"
                    item_divs = group_div.find_all('div', {
                        'role': 'treeitem', 'class': '_p13n-zg-nav-tree-all_style_zg-browse-item__1rdKf'})
                    
                    # Loop through item divs to extract names and URLs
                    for item_div in item_divs:
                        # Find the anchor tag within the div
                        anchor_tag = item_div.find('a')
                        if anchor_tag:
                            category_name = anchor_tag.text.strip()  # Extract category name
                            # Append base URL to category URL
                            category_url = base_url + anchor_tag['href']
                            url_data.append(
                                {'name': category_name, 'url': category_url})
                            # worksheet.cell(row=row, column=2).value=category_url
                            # worksheet.cell(row=row, column=1).value=category_name

                            row += 1
                    # workbook.save('./amazon_categories_all.xlsx')

                            # Print the data on the screen
                            # print(f'Name: {category_name}\nURL: {category_url}\n')

                break  # Break out of the loop if successful

            except TimeoutException:
                print("Timed out waiting for page to load, retrying...")

def sub_scrape_data_from_url(urls, array_to_append, is_worksheet):
    row = 1
    for url in urls:
        if url == '':
            continue
        for _ in range(max_retries):
            try:
                # Navigate to the URL using Selenium
                driver.get(url['url'])

                # Wait for the page to load and find the relevant elements
                WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                    (By.CLASS_NAME, '_p13n-zg-nav-tree-all_style_zg-browse-group__88fbz')))

                # Parse the HTML content of the page using BeautifulSoup
                soup = BeautifulSoup(driver.page_source, 'html.parser')

                # Find all div elements with role="group" and class="_p13n-zg-nav-tree-all_style_zg-browse-group__88fbz"
                group_divs = soup.find_all(
                    'div', {'class': '_p13n-zg-nav-tree-all_style_zg-browse-group__88fbz'})

                # print(len(group_divs))
                # Loop through each group div and extract names and URLs
                for index, group_div in enumerate(group_divs):
                    # Find all div elements with role="treeitem" and class="_p13n-zg-nav-tree-all_style_zg-browse-item__1rdKf"
                    categories = []
                    if index == 0:
                        item_divs = group_div.find_all('div', {
                            'role': 'treeitem', 'class': '_p13n-zg-nav-tree-all_style_zg-browse-item__1rdKf'})

                        # Loop through item divs to extract names and URLs
                        for item_div in item_divs[:2]:
                            # Find the anchor tag within the div
                            anchor_tag = item_div.find('a')
                            span_tag = item_div.find('span')

                            if span_tag:
                                categories.append(
                                    {'name': span_tag.text.strip(), 'url': ''})

                            if anchor_tag:
                                category_name = anchor_tag.text.strip()  # Extract category name
                                # Append base URL to category URL
                                category_url = base_url + anchor_tag['href']
                                categories.append(
                                    {'name': category_name, 'url': category_url})

                        # print(categories)
                        for category in categories:
                            array_to_append.append(
                                {"name": category['name'], "url": category['url']})
                            if is_worksheet != '':
                                worksheet.cell(
                                    row=row, column=2).value = category['url']
                                worksheet.cell(
                                    row=row, column=1).value = category['name']
                                row += 1
                    if index == 2:
                        item_divs = group_div.find_all('div', {
                            'role': 'treeitem', 'class': '_p13n-zg-nav-tree-all_style_zg-browse-item__1rdKf'})

                        # Loop through item divs to extract names and URLs
                        for item_div in item_divs:
                            # Find the anchor tag within the div
                            anchor_tag = item_div.find('a')
                            span_tag = item_div.find('span')

                            if span_tag:
                                categories.append(
                                    {'name': span_tag.text.strip(), 'url': ''})

                            if anchor_tag:
                                category_name = anchor_tag.text.strip()  # Extract category name
                                # Append base URL to category URL
                                category_url = base_url + anchor_tag['href']
                                categories.append(
                                    {'name': category_name, 'url': category_url})
                        # print(categories)
                        for category in categories:
                            array_to_append.append(
                                {"name": category['name'], "url": category['url']})
                            if is_worksheet != '':
                                worksheet.cell(
                                    row=row, column=2).value = category['url']
                                worksheet.cell(
                                    row=row, column=1).value = category['name']
                                row += 1
                    workbook.save('./amazon_categories_all.xlsx')
                break

            except TimeoutException:
                print("Timed out waiting for page to load, retrying...")


main_scrape_data_from_url(url)
for url_d in url_data:
    print(url_d)
print('Now the First Categories data will be started to be add in Worksheet EXCEL FILE')
sub_scrape_data_from_url(url_data, sub_url_data, 'Add')


# Initialize row counter
row = 1

# Function to handle HTTP request errors with retries


def get_url_with_retry(url, max_retries=3):
    for i in range(max_retries):
        try:
            # sleep(5)
            response = driver.get(url)
            # response.raise_for_status()  # Raise an exception for 4xx and 5xx status codes
            print('Successfull')
            return response
        except requests.exceptions.RequestException as e:
            print(f"Request failed. Retrying... ({i+1}/{max_retries})")
    raise Exception("Failed to retrieve the page after multiple retries")

# Define a function to wait for an element with retries


def wait_for_element(by, value, retries=3):
    for _ in range(retries):
        try:
            return WebDriverWait(driver, 30).until(EC.presence_of_element_located((by, value)))
        except TimeoutException:
            print("Timeout occurred. Retrying...")
    raise TimeoutException(f"Timed out waiting for element: {by}={value}")


# Iterate through the URLs and their names to create rows for headings and store URLs
worksheet.cell(row=row, column=1, value='Name')
worksheet.cell(row=row, column=2, value='URLs')
row += 1

print('Now the Second Categories data will be started to be add in Worksheet EXCEL FILE')

for url_info in sub_url_data:
    url_name = url_info['name']
    url = url_info['url']

    print(f"Processing {url_name} ({url})")  # Add a print statement to debug

    # Add URL name and URL as headings in separate rows

    worksheet.cell(row=row, column=1, value=url_name)
    worksheet.cell(row=row, column=2, value=url)
    row += 1

    # Get the page content with retries
    try:
        response = get_url_with_retry(url)
    except Exception as e:
        print(f"Failed to retrieve the page ({url_name}). Error: {str(e)}")
        continue

    # Navigate to the URL using Selenium
    driver.get(url)

    # Wait for the entire page to load
    wait_for_element(By.TAG_NAME, 'body')

    # Get the page source after it's fully loaded
    page_source = driver.page_source

    # Parse the page source with BeautifulSoup
    soup = BeautifulSoup(page_source, 'html.parser')

    # Find all div elements with role="group" and class="_p13n-zg-nav-tree-all_style_zg-browse-group__88fbz"
    group_divs = soup.find_all(
        'div', {'class': '_p13n-zg-nav-tree-all_style_zg-browse-group__88fbz'})

    # print(len(group_divs))
    # Loop through each group div and extract names and URLs
    for index, group_div in enumerate(group_divs):
        # Find all div elements with role="treeitem" and class="_p13n-zg-nav-tree-all_style_zg-browse-item__1rdKf"
        categories = []
        if index == 0:
            item_divs = group_div.find_all('div', {
                'role': 'treeitem', 'class': '_p13n-zg-nav-tree-all_style_zg-browse-item__1rdKf'})

            # Loop through item divs to extract names and URLs
            for item_div in item_divs[:2]:
                # Find the anchor tag within the div
                anchor_tag = item_div.find('a')
                span_tag = item_div.find('span')

                if span_tag:
                    categories.append(
                        {'name': span_tag.text.strip(), 'url': ''})

                if anchor_tag:
                    category_name = anchor_tag.text.strip()  # Extract category name
                    # Append base URL to category URL
                    category_url = base_url + anchor_tag['href']
                    categories.append(
                        {'name': category_name, 'url': category_url})

            # print(categories)
            for category in categories:
                worksheet.cell(row=row, column=1, value=category['name'])
                worksheet.cell(row=row, column=2, value=category['url'])
                row += 1
        if index == 2:
            item_divs = group_div.find_all('div', {
                'role': 'treeitem', 'class': '_p13n-zg-nav-tree-all_style_zg-browse-item__1rdKf'})

            # Loop through item divs to extract names and URLs
            for item_div in item_divs:
                # Find the anchor tag within the div
                anchor_tag = item_div.find('a')
                span_tag = item_div.find('span')

                if span_tag:
                    categories.append(
                        {'name': span_tag.text.strip(), 'url': ''})

                if anchor_tag:
                    category_name = anchor_tag.text.strip()  # Extract category name
                    # Append base URL to category URL
                    category_url = base_url + anchor_tag['href']
                    categories.append(
                        {'name': category_name, 'url': category_url})
            # print(categories)
            for category in categories:
                worksheet.cell(row=row, column=1, value=category['name'])
                worksheet.cell(row=row, column=2, value=category['url'])
                row += 1
    # Save the Excel workbook with all data
    workbook.save('amazon_categories_all.xlsx')

# Close the WebDriver
driver.quit()
