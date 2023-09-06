from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from bs4 import BeautifulSoup
import re

# Load the Excel workbook
workbook = openpyxl.load_workbook(r'./Macro_file.xlsx')
worksheet = workbook.active  # Assumes the data is in the active sheet

# Define the path to the CRX file
extension_path = './amazoncrxextension.crx'

# Set up Chrome WebDriver with the extension
chrome_options = webdriver.ChromeOptions()
chrome_options.add_extension(extension_path)

# Create the WebDriver with ChromeOptions
driver = webdriver.Chrome(options=chrome_options)

# Define the Amazon URL

# Set up WebDriverWait
wait = WebDriverWait(driver, 15)

# Loop through each row in the Excel sheet
for row_index in range(2, worksheet.max_row + 1):
    asin = worksheet.cell(row=row_index, column=1).value  # Get the ASIN from the current row

    if asin is not None:
        # Navigate to Amazon
        driver.get(f'https://www.amazon.de/dp/{asin}')

        try:
            # Check for the presence of the "sp-cc-rejectall-link" element
            reject_button = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID, 'sp-cc-rejectall-link')))
    
            # Click the "Continue without accepting" link
            reject_button.click()
        except Exception as e:
            pass
            # print(f"Could not click on 'sp-cc-rejectall-link': {e}")

        try:

            # Capture the entire HTML of the page and parse it with BeautifulSoup
            page_html = driver.page_source
            soup = BeautifulSoup(page_html, 'html.parser')

            # Find the product title element by its ID
            product_title_element = soup.find("span", {"id": "productTitle"})   
            product_brand_element = soup.find("a", {"id": "bylineInfo"})
            product_sales_element = soup.find("span", {"class": "dpWidgetSoldUnitsLabelSum"})
            product_number_of_reviews_element = soup.find("span", {"id": "acrCustomerReviewText"})
            product_avg_review_element = soup.find("div", {"id": "averageCustomerReviews"})  # Updated selector
            first_anchor_tag = soup.find('div', id='wayfinding-breadcrumbs_feature_div').find('a')

            # Extract the text from the first anchor tag
            first_anchor_text = first_anchor_tag.text.strip()
            price = 'Price not found'
            rank_number = 'Best Sellers Rank not found'
            try:
                product_price_element = soup.find("span", {"id": "price_inside_buybox"})
                price = product_price_element.text.strip()
            except Exception as e:
                print("Price not found")

            # Extract the average rating sentence
            if product_avg_review_element:
                avg_rating_sentence = product_avg_review_element.find("span", {"class": "a-declarative"}).text.strip()
            else:
                avg_rating_sentence = 'Average Rating not found'

            # Find the Best Sellers Rank element
            links = soup.find_all('th', {"class": "a-color-secondary a-size-base prodDetSectionEntry"})
            bestSellerTD = ''
            avgReviewTD = ''
            for link in links:
                if link.find(string=re.compile("Best Sellers Rank")):
                    bestSellerTD = link
                if link.find(string=re.compile("Customer Reviews")):
                    avgReviewTD = link

            # Check if the <th> element was found
            if bestSellerTD:
                # Find the next <td> element which contains the actual data
                best_seller_td_element = bestSellerTD.find_next_sibling("td")
                avg_review_td_element = avgReviewTD.find_next_sibling("td")

                # Extract the text from the <td> element
                rank_info = best_seller_td_element.get_text(strip=True)
                rank_info_text = best_seller_td_element.get_text(strip=True)
                match = re.search(r'(\d+(?:,\d+)*)', rank_info)

                if match:
                    rank_number = match.group(1)
                else:
                    rank_number = 'Best Sellers Rank not found'
            else:
                print("Best Sellers Rank not found on the page.")

            # Extract the content of the product title and brand elements
            if product_title_element and product_brand_element:
                title = product_title_element.text.strip()
                full_brand = product_brand_element.text.strip()
                sales_number = product_sales_element.text.strip()
                reviewNumbers = product_number_of_reviews_element.text.strip()
                avgReview = avg_rating_sentence[:3]

                # Extract the brand name as you described
                brand_parts = full_brand.split()
                if len(brand_parts) > 2:  # Check if there are at least 3 words in the brand
                    brand = ' '.join(brand_parts[2:-1])  # Omit the first two words and the last word
                else:
                    brand = full_brand  # Use the original brand if it doesn't match the expected format
                sales_element_text = product_sales_element.text.strip()
                match = re.search(r'∑\s*(\d+)\s*\(.*\)', product_sales_element.text.strip())

                if match:
                    sales_number = match.group(1)
                else:
                    sales_number = 'Sales Number not found'
                    print(f"sales_number match: {match}")

                print(f"Title: {title}")
                print(f"Brand: {brand}")
                print(f"Price: {price}")
                print(f"Best Sellers Rank: {rank_number}")
                print(f"Review Number: {reviewNumbers}")
                print(f"Sales Number: {sales_number}")
                print(f"Average Review: {avgReview}")
                print(f"Rank Category: {first_anchor_text}")
                
                # Update the Excel sheet with the retrieved data
                worksheet.cell(row=row_index, column=2, value=brand)
                worksheet.cell(row=row_index, column=3, value=title)
                worksheet.cell(row=row_index, column=4, value=price)
                worksheet.cell(row=row_index, column=5, value=rank_number)
                worksheet.cell(row=row_index, column=6, value=reviewNumbers)
                worksheet.cell(row=row_index, column=7, value=sales_number)
                worksheet.cell(row=row_index, column=8, value=avgReview)
                worksheet.cell(row=row_index, column=9, value=first_anchor_text)

                # Save the Excel workbook
                workbook.save(r'./Macro_file.xlsx')
            else:
                print(f"ASIN {asin} is sponsored. Skipping to the next ASIN.")

        except Exception as e:
            print(f"ASIN {asin} not found or encountered an error: {e}. Skipping to the next ASIN.")
    else:
        print(f"ASIN is None for row {row_index}. Skipping to the next row.")

# Close the WebDriver
driver.quit()
